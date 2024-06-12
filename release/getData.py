import openpyxl
from calculate import *
import io
# from datetime import datetime
def init(file, multi_, lc_ = 0.1):
    in_mem_file = None
    with open(file, "rb") as f:
        in_mem_file = io.BytesIO(f.read())

    wb = openpyxl.load_workbook(in_mem_file)
    # wb = openpyxl.load_workbook(file)
    global ws
    global multi
    global results
    global lc
    ws = wb.active
    multi = multi_
    lc = lc_

    # try to get all data from sheets
    results = dict()
    for col in ws.iter_cols():
        title = col[0].value
        if title == None: continue
        values = [cell.value for cell in col[1:]]
        results[title] = values

def getData() -> dict:
    global multi
    global results
    global lc
    data = dict()
    if multi:
        for k, v in results.items():
            data[k] = [Data(v, lc)]
    else:
        for k, v in results.items():
            data[k] = []
            for i in v:
                data[k].append(Data([i], lc))
    return data

def saveResults(result: list, file: str):
    file = file.split('/')
    # time = str(datetime.now())[:-7]
    path = f"{'/'.join(file[:-1])}/{file[-1]}計算結果.txt".replace("\\", "/")
    with open(path, encoding = "utf-8", mode = "w") as file:
        for i in result:
            file.write(str(i) + "\n")
    return path

    



# def getCol(col: int, lc: float = 0.1): 
#     # multi: 是一個數據的多次測量(有 a) 還是很多組不同的數據(沒有 a)
#     # return 一整個 column 從第 2 行開始(第一行標題)
#     global ws
#     global multi
#     if multi:
#         x = []
#         for row in ws.iter_rows(min_col = col, max_col = col, min_row = 2):
#             for cell in row:
#                 # print(cell.value)
#                 if cell.value != None and type(cell.value) != str: 
#                     x.append(cell.value)
#         # print("Get column", col, ":", x)
#         return Data(x, lc)
#     else:
#         x = []
#         for row in ws.iter_rows(min_col = col, max_col = col, min_row = 2):
#             for cell in row:
#                 # print(cell.value)
#                 if cell.value != None and type(cell.value) != str: 
#                     x.append(Data([cell.value], lc))
#         # print("Get column", col, ":", x)
#         return x


